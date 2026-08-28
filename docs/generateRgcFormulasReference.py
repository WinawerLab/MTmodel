#!/usr/bin/env python3
"""Generate RGC formulas reference (XLSX + HTML) for the live MTmodel front-end.

Outputs to ../Documents/ (Computation MT Model/Documents/) by default.
Run:  python3 docs/generateRgcFormulasReference.py

Updated 2026-08-26 for: two-preset API, lagged biological front-end, mtMix,
fourPop removal, docs/MODEL_AND_LESIONS.md as authority.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
OUT_DIR = REPO.parent / "Documents"


def style_header(ws, row=1):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(len(str(c.value or "")) for c in col) + 2, max_width)
        ws.column_dimensions[letter].width = width


def add_sheet(ws, headers, rows):
    ws.append(headers)
    style_header(ws)
    for row in rows:
        ws.append(row)
    autosize(ws)


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    # --- Overview ---
    ws = wb.create_sheet("Overview")
    add_sheet(
        ws,
        ["Topic", "Summary"],
        [
            ["Last updated", "2026-08-26"],
            ["Authority", "docs/MODEL_AND_LESIONS.md, pars/shPars.m, shClassV1Basis.m"],
            [
                "Two ways to run",
                "shPars or shPars('derivative') — SH basis, exact legacy; "
                "shPars('lagged') — biological 16-class front-end + two-stream MT (mtMix)",
            ],
            [
                "Code path",
                "Single class-based path: pars.rgc.classes → shClassV1Basis → "
                "shModelV1LinearFromClasses. No separate fourPop / shModelRgcPopulation.",
            ],
            [
                "Driving question",
                "Can RGC damage explain (a) slower VEP and (b) worse motion-defined form at slow speeds?",
            ],
            [
                "Physical scale",
                "1 px = 0.430 deg; 1 frame = 26.9 ms (37.2 fps); 1 px/frame = 16 deg/s (shModelUnits.m)",
            ],
            [
                "Retired",
                "fourPop preset, midgetParasol offset+quadrature, shModelRgcPopulation — deleted or in explore/_archive/",
            ],
        ],
    )

    # --- Pipeline ---
    ws = wb.create_sheet("Pipeline")
    add_sheet(
        ws,
        ["Step", "Function / field", "Output"],
        [
            ["1", "Stimulus I(x,y,t)", "[Y×X×T] movie in [0,1]"],
            ["2", "Per-class channel", "localClassChannel in shClassV1Basis.m"],
            ["2a", "Frame mean subtract (DoG classes)", "I - mean(I) per frame"],
            ["2b", "DoG spatial RF", "center - w_s × surround (separable Gaussian)"],
            ["2c", "ON/OFF rectify", "max(0,·) or max(0,-·) per class.rectify"],
            ["2d", "Causal temporal conv", "convn(m, k, 'full') → trim to T frames"],
            ["2e", "Class gain", "× classes(i).gain (lesion hook)"],
            ["3", "Impairment maps", "shApplyRgcImpairment — amp × map, per-pixel delay"],
            ["4", "V1 derivative features", "shClassV1Basis — valid correlation with SF orders 0–3"],
            ["5", "Combine", "steer (derivative) or weights (lagged)"],
            ["6", "V1 complex → MT", "standard SH normalization; optional mtMix (§ MT_streams)"],
        ],
    )

    # --- Formulas ---
    ws = wb.create_sheet("Formulas")
    add_sheet(
        ws,
        ["Name", "Formula", "Notes"],
        [
            ["Gaussian 1D", "g[n] = exp(-n²/2σ²) / Σ exp(...), support ±ceil(3σ)", "mkGaussianFilter; σ=-1 → identity"],
            ["Separable 2D blur", "(G*I)(x,y) = Σ g[x-x'] Σ g[y-y'] I(x',y',t)", "convn, boundary 'same'"],
            ["DoG RF", "R = G_c*I - w_s·G_s*I", "localDoG in shClassV1Basis"],
            ["Biological contrast", "I_c = I - mean_{x,y}(I) per frame", "Then DoG on I_c; not pars.rgc.onOffSignSplit"],
            ["ON half-wave", "m = max(0, R)", "rectify = onHalf"],
            ["OFF half-wave", "m = max(0, -R)", "rectify = offHalf"],
            ["Bi-gamma kernel", "k(t) ∝ (t/τ1)^n e^{-t/τ1} - w(t/τ2)^n e^{-t/τ2}", "causal t=0…L-1; normalized by max|k|"],
            ["Lag copy", "k_d = [zeros(d,1); k(0:L-1)]", "Healthy tiling; NOT impairment delay"],
            ["Impairment amp", "m' = A(x,y)·m", "pars.rgc.impairmentAmplitudeMap"],
            ["Impairment delay", "m'(t) = m(t - D(x,y))", "integer frames; shApplyRgcImpairment"],
            ["Class amplitude lesion", "classes(i).gain = g", "multiplicative on channel output"],
            ["Class delay lesion", "prepend zeros to temporalKernel", "causal shift of filter"],
            ["V1 steer", "analytic shSwts weights", "derivative preset, 10 features total"],
            ["V1 weights", "W @ features → 28 neurons", "ridge fit; lagged preset"],
            ["mtMix", "popMT = (1-α)·MT(popA) + α·delay(MT(popB), d)", "after V1 normalization; shModelV1ComplexForMt"],
        ],
    )

    # --- Biological ON/OFF ---
    ws = wb.create_sheet("Biological_channel")
    add_sheet(
        ws,
        ["Item", "Value", "Why"],
        [
            [
                "Polarity mechanism",
                "Separate ON and OFF classes on same DoG RF",
                "Not global onOffSignSplit; each class has rectify onHalf/offHalf",
            ],
            ["Contrast reference", "Global frame mean subtract before DoG", "localClassChannel in shClassV1Basis"],
            ["readoutOffset", "[0 0] for lagged preset", "Spatial offset DS mechanism retired 2026-07-12"],
            ["ON quadrature", "None", "Retired with offset preset"],
        ],
    )

    # --- Temporal ---
    ws = wb.create_sheet("Temporal")
    add_sheet(
        ws,
        ["Parameter", "Parasol (fast)", "Midget (slow)", "Units"],
        [
            ["τ1", "0.6", "2.0", "frames"],
            ["τ2", "1.2", "4.0", "frames"],
            ["w (neg weight)", "0.45", "0.15", "—"],
            ["n (gamma order)", "2", "2", "—"],
            ["L (length)", "24", "24", "frames"],
            ["Peak (approx)", "~1 frame / 27 ms", "~4 frames / 107 ms", "see TODO §6 Kling calibration"],
            ["Default lags", "0, 1, 2, 3", "same", "frames → 16 classes"],
        ],
    )

    # --- Spatial RF ---
    ws = wb.create_sheet("Spatial_RF")
    add_sheet(
        ws,
        ["Parameter", "Midget", "Parasol", "Notes"],
        [
            ["centerSigma", "0.8 px", "1.6 px", "≈0.34 deg vs 0.69 deg at model scale"],
            ["surroundSigma", "2.0 px", "4.0 px", "—"],
            ["surroundWeight", "0.25", "0.25", "integrated surround ~12–13% of centre"],
            ["Features per class", "10", "10", "readoutOrders [0 1 2 3] → all derivative orders"],
            ["Total features", "160", "160", "16 classes × 10"],
        ],
    )

    # --- pars.rgc ---
    ws = wb.create_sheet("pars_rgc")
    add_sheet(
        ws,
        ["Field", "Default / values", "Role"],
        [
            ["enabled", "1", "RGC front-end on (legacy oracle: set 0)"],
            ["mode", "'derivative' or 'custom'", "Set by shPars; custom keeps hand-built classes"],
            ["classes", "from preset", "array of shRgcClass structs"],
            ["combine", "'steer' | 'weights'", "derivative → steer; lagged → weights"],
            ["classesMode", "matches preset", "do not hand-edit without mode='custom'"],
            ["v1Weights", "[] or fitted W", "28×160 for lagged; unused for derivative"],
            ["derivative.channelGain", "[1 1 1 1]", "per temporal-order lesion on derivative preset"],
            ["impairmentEnabled", "0", "spatial amp/delay maps"],
            ["impairmentAmplitudeMap", "Y×X", "multiplicative per pixel"],
            ["impairmentDelayMap", "Y×X integer", "frames delay per pixel"],
            ["mtMix.weightsA", "28×160 parasol mask", "stream A (4B→MT); lagged preset only"],
            ["mtMix.alpha", "0.10", "weight on slow mixed stream B"],
            ["mtMix.delay", "0", "integer frames on stream B (V2 detour)"],
        ],
    )

    # --- Class fields ---
    ws = wb.create_sheet("Class_fields")
    add_sheet(
        ws,
        ["Field", "Type", "Default", "Meaning"],
        [
            ["name", "char", "—", "e.g. parasolOn_lag2"],
            ["temporalKernel", "column L×1", "—", "causal filter; may include leading zeros for lag"],
            ["spatialRF", "struct or []", "[]", "[] = delta; else DoG sigmas"],
            ["rectify", "char", "'none'", "none | onHalf | offHalf"],
            ["readoutOrders", "vector", "[0 1 2 3]", "which SF derivative orders fed"],
            ["readoutOffset", "[dy dx]", "[0 0]", "spatial shift at V1 read-out"],
            ["gain", "scalar", "1", "lesion: multiply channel response"],
        ],
    )

    # --- Presets ---
    ws = wb.create_sheet("Presets")
    add_sheet(
        ws,
        ["Call", "Front-end", "combine", "MT", "Healthy fidelity", "Notes"],
        [
            [
                "shPars('derivative')",
                "4 classes, SH temporal orders 0–3, delta RF",
                "steer",
                "single stream",
                "err = 0 vs legacy",
                "Oracle; machine-precision reference",
            ],
            [
                "shPars('lagged')",
                "16 classes: midget/parasol × ON/OFF × lags 0–3",
                "weights",
                "mtMix on (α=0.10)",
                "pooled r ≈ 0.984 (see caveats §4.2 report)",
                "Biological lesion parameterization; M/P at MT",
            ],
            [
                "Clear mtMix",
                "same as lagged",
                "weights",
                "single mixed stream",
                "—",
                "Reproduces pre-2026-08-14 MT; midget-dominated",
            ],
        ],
    )

    # --- MT streams ---
    ws = wb.create_sheet("MT_streams")
    add_sheet(
        ws,
        ["Stream", "Weights", "Anatomy", "Role"],
        [
            [
                "A (popA)",
                "parasol-masked W_A from shFitClassV1Weights + feature mask ^parasol",
                "4B → MT (76% spiny stellate, 4Cα)",
                "Dominant fast magno drive; (1-α) of mixture",
            ],
            [
                "B (popB)",
                "existing pars.rgc.v1Weights (mixed M+P)",
                "→ V2 thick stripes → MT",
                "Slow minority; α=0.10 default",
            ],
            [
                "Mixture",
                "popMT = (1-α)·MT(popA) + α·delay(popB,d)",
                "Nassi & Callaway 2006/2007",
                "Formed after V1 normalization (separate pools)",
            ],
            [
                "Calibration",
                "Maunsell et al. 1990 knockouts",
                "M block strong; P block weak for typical unit",
                "α must stay 0.05–0.10",
            ],
        ],
    )

    # --- Lesions ---
    ws = wb.create_sheet("Lesions")
    add_sheet(
        ws,
        ["Type", "Mechanism", "Example", "Interpretation"],
        [
            ["Uniform amplitude", "classes(i).gain = 0.5 all", "50% signal loss", "Optic nerve global"],
            ["Parasol amplitude", "gain=0.3 on name contains 'parasol'", "70% M reduction", "Needs mtMix for biology"],
            ["Midget amplitude", "gain=0.3 on 'midget'", "70% P reduction", "Speed-graded effect at MT §4.5"],
            ["Uniform delay", "prepend 2 frames to all kernels", "+2 frame conduction", "Invisible to steady-state tuning"],
            ["ON-only delay", "prepend 1 frame on onHalf classes", "pathway-specific", "Biological preset"],
            ["Spatial amp map", "impairmentAmplitudeMap", "random U(0.3,0.7)", "heterogeneous SNR"],
            ["Spatial delay map", "impairmentDelayMap", "per-pixel {0,1,2,3}", "delay_random — largest coherence hit"],
            ["Weights fixed", "no refit after lesion", "—", "isolates RGC effect from cortical adaptation"],
        ],
    )

    # --- Units ---
    ws = wb.create_sheet("Units")
    add_sheet(
        ws,
        ["Quantity", "Model value", "Source"],
        [
            ["deg per pixel", "0.430", "SH Appendix I via 0.2148 cyc/px ↔ 0.5 cyc/deg"],
            ["pixels per degree", "2.33", "shModelUnits.m"],
            ["ms per frame", "26.9", "37.2 fps"],
            ["deg/s per px/frame", "16", "product of above"],
            ["MT speed shells (nominal)", "0, 1, 6 px/frame", "0, 16, 96 deg/s"],
            ["Clinical low-speed band", "0.6–9.6 deg/s", "below MT slowest moving unit"],
        ],
    )

    # --- Convolution ---
    ws = wb.create_sheet("Convolution")
    add_sheet(
        ws,
        ["Operation", "MATLAB", "Boundary / trim"],
        [
            ["Spatial separable", "convn twice (row, col)", "'same'"],
            ["Temporal causal", "convn(m, k, 'full')", "keep frames 1:T"],
            ["V1 valid corr", "shValidCorrDn3", "drops invalid borders"],
            ["Delay map", "per-pixel integer shift along t", "shApplyRgcImpairment"],
        ],
    )

    return wb


def build_html():
    return """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>MTmodel — RGC Formulas Reference (2026-08-26)</title>
<style>
  body { font-family: Georgia, 'Times New Roman', serif; max-width: 920px; margin: 2em auto; padding: 0 1.5em; line-height: 1.55; color: #222; }
  h1 { font-size: 1.6em; border-bottom: 2px solid #4472C4; padding-bottom: 0.3em; }
  h2 { font-size: 1.25em; color: #4472C4; margin-top: 1.8em; }
  h3 { font-size: 1.05em; margin-top: 1.2em; }
  table { border-collapse: collapse; width: 100%; margin: 1em 0; font-size: 0.92em; }
  th, td { border: 1px solid #ccc; padding: 8px 10px; text-align: left; vertical-align: top; }
  th { background: #4472C4; color: white; }
  tr:nth-child(even) { background: #f7f9fc; }
  code { background: #f0f0f0; padding: 1px 4px; font-size: 0.9em; }
  .formula { background: #f9f9f9; border-left: 4px solid #4472C4; padding: 0.8em 1em; margin: 1em 0; font-family: 'Courier New', monospace; white-space: pre-wrap; }
  .note { background: #fff8e6; border: 1px solid #e6d9a8; padding: 0.8em 1em; margin: 1em 0; }
  @media print { body { max-width: none; margin: 1cm; } h2 { page-break-after: avoid; } table { page-break-inside: avoid; } }
</style>
</head>
<body>
<h1>MTmodel — RGC Formulas Reference</h1>
<p><strong>Updated 2026-08-26.</strong> Companion <strong>.xlsx</strong> has sortable tables (13 sheets).
Authority: <code>docs/MODEL_AND_LESIONS.md</code>, <code>pars/shPars.m</code>, <code>shClassV1Basis.m</code>.</p>

<div class="note"><strong>Two ways to run:</strong> <code>shPars('derivative')</code> (exact SH) and
<code>shPars('lagged')</code> (biological 16-class front-end + two-stream MT). The fourPop preset and
<code>shModelRgcPopulation</code> path are <strong>removed</strong>.</div>

<h2>1. Pipeline</h2>
<p>Stimulus I(x,y,t) → per-class channels (<code>shClassV1Basis</code>) → V1 derivative features →
combine (steer or weights) → V1 complex → MT (optional <code>mtMix</code>).</p>

<h2>2. Spatial: DoG receptive field</h2>
<div class="formula">Frame contrast:  I_c(x,y,t) = I(x,y,t) - mean_{x,y} I(·,·,t)
DoG:             R = G_c * I_c - w_s · G_s * I_c
ON/OFF:          m_ON = max(0, R)   or   m_OFF = max(0, -R)   [per class]</div>
<table>
<tr><th></th><th>Midget</th><th>Parasol</th></tr>
<tr><td>centerSigma</td><td>0.8 px</td><td>1.6 px</td></tr>
<tr><td>surroundSigma</td><td>2.0 px</td><td>4.0 px</td></tr>
<tr><td>surroundWeight</td><td>0.25</td><td>0.25</td></tr>
</table>

<h2>3. Temporal: bi-gamma + lags</h2>
<div class="formula">k(t) = (1/P)[ (t/τ₁)^n exp(-t/τ₁) - w (t/τ₂)^n exp(-t/τ₂) ],  t = 0…23
Lag d:  k_d = [zeros(d,1); k]     (healthy tiling — NOT impairment)
Channel: convn(m, k, 'full') → frames 1:T</div>
<table>
<tr><th></th><th>Parasol</th><th>Midget</th></tr>
<tr><td>τ₁ / τ₂</td><td>0.6 / 1.2 fr</td><td>2.0 / 4.0 fr</td></tr>
<tr><td>w</td><td>0.45</td><td>0.15</td></tr>
</table>
<p>Default lags {0,1,2,3} → 16 classes × 10 read-outs = <strong>160 features</strong>.</p>

<h2>4. V1 read-out</h2>
<p><strong>derivative:</strong> <code>combine = 'steer'</code> — 4 classes, analytic SH steering, 10 features, exact legacy.</p>
<p><strong>lagged:</strong> <code>combine = 'weights'</code> — ridge-fit W (28×160), cached in <code>pars/</code>.</p>

<h2>5. Two-stream MT (lagged preset)</h2>
<div class="formula">popMT = (1 - α) · MT(popA)  +  α · delay(MT(popB), d)

popA  weightsA = parasol-masked fit  (stream "4B→MT", magno)
popB  pars.rgc.v1Weights            (stream "→V2→MT", mixed M+P)
Default: α = 0.10, d = 0</div>
<p>Mixture is formed <strong>after</strong> V1 normalization. Clear <code>pars.rgc.mtMix</code> for single-stream (pre-2026-08-14) behaviour.</p>

<h2>6. Lesions</h2>
<table>
<tr><th>Axis</th><th>Class-level</th><th>Spatial map</th></tr>
<tr><td>Amplitude</td><td>classes(i).gain</td><td>impairmentAmplitudeMap</td></tr>
<tr><td>Delay</td><td>prepend zeros to temporalKernel</td><td>impairmentDelayMap (integer frames/pixel)</td></tr>
</table>
<p>Applied in <code>shApplyRgcImpairment</code> after channel formation. V1 weights are <strong>not</strong> refit.</p>

<h2>7. Physical units (shModelUnits)</h2>
<table>
<tr><th>Quantity</th><th>Value</th></tr>
<tr><td>1 pixel</td><td>0.430 deg (2.33 px/deg)</td></tr>
<tr><td>1 frame</td><td>26.9 ms (37.2 fps)</td></tr>
<tr><td>1 px/frame</td><td>16 deg/s</td></tr>
</table>

<h2>8. Presets (live only)</h2>
<table>
<tr><th>Call</th><th>Classes</th><th>MT</th><th>Fidelity</th></tr>
<tr><td>shPars('derivative')</td><td>4 SH orders</td><td>single</td><td>0 vs legacy</td></tr>
<tr><td>shPars('lagged')</td><td>16 midget/parasol lagged</td><td>mtMix α=0.10</td><td>~0.984 pooled r (see report §4.2)</td></tr>
</table>

<p style="margin-top:2em;font-size:0.85em;color:#666;">Regenerate: <code>python3 MTmodel/docs/generateRgcFormulasReference.py</code> → Computation MT Model/Documents/</p>
</body>
</html>
"""


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = OUT_DIR / "RGC_formulas_reference.xlsx"
    html_path = OUT_DIR / "RGC_formulas_reference.html"

    wb = build_workbook()
    wb.save(xlsx_path)
    html_path.write_text(build_html(), encoding="utf-8")

    print(f"Wrote {xlsx_path}")
    print(f"Wrote {html_path}")
    print("PDF: open HTML in browser → Print → Save as PDF, or run headless Chrome.")


if __name__ == "__main__":
    main()
